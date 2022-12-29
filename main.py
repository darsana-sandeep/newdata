import threading
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook


#create a workbook
wb = Workbook()
ws = wb.active
wb.save('Eexcs.xlsx')
import random

ws = wb.active
n=2
while n <=20:
        x=random.randint(10,100)

        c1 = ws['A1']
        c1.value = "input"
        c2 = ws['A' + str(n)]
        c2.value = x

        c4 = ws['B1']
        c4.value = "Freq"

        y=random.randint(2,100)
        c3 = ws['B' + str(n)]
        c3.value = y
        wb.save('Eexcs.xlsx')

        n= n+1

# df_days_calories = pd.DataFrame(
#         {'input': x,
#          'Freq': y
#          })
#
# ax = plt.gca()
#
# # use plot() method on the dataframe
# df_days_calories.plot(x='input', y='Freq', ax=ax)

df = pd.read_excel('DATA SET & RULES - MOTOR TESTINdemo.xlsx',engine='openpyxl') # only for xlsx files




